#!/usr/bin/env python
# Script to incorporate the AMRF+ & CARD reference lists into one, then combines AMRF/CARD & RESF/BN combi-lists into one
# 
# Importing the required libraries
from openpyxl import load_workbook
import xlsxwriter

# Saving the paths to the files in an object
AMRF_list = "/home/guest/BIT11_Traineeship/Ecoli_AMR/Reference_lists/AMRF_reflist.xlsx"
CARD_list = "/home/guest/BIT11_Traineeship/Ecoli_AMR/Reference_lists/CARD_reflist.xlsx" 
outfile = "/home/guest/BIT11_Traineeship/Ecoli_AMR/Reference_lists/combi_reflist_AMRF_CARD.xlsx" 

# Loading input-& output-file
AMRF_wb = load_workbook(AMRF_list)
CARD_wb = load_workbook(CARD_list)

############ AMRF + CARD ############ 
# STEP 1 : Incorporating data from AMRF- & CARD-reference lists in a dictionary
###########################################################################################################################

# Initializing dictionaries to store the data
combi_dict = {} # Will contain AMRF data first, then CARD data will be added later in the script, eventually it will also have RESF and BN data
CARD_dict = {}

# Function to create a dictionary from a workbook's first sheet
def create_dict(wb,dict):
    ws = wb.active
    for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        key = row[0]
        value = row[1]
        
        if value is not None:
            value = value.lower()
            antibiotics = [antibiotic.strip() for antibiotic in value.split(',')]  # Split the string into individual antibiotics and strip whitespace
            dict[key] = antibiotics
        else:
            dict[key] = [] # CARD genes (e.g. acrD) sometimes don't have an antibiotic linked, this will correcty leave the dictionary value empty in that case

# Storing the data from combi_wb & BN_wb in a dictionary each
create_dict(AMRF_wb, combi_dict)
create_dict(CARD_wb, CARD_dict)

# Combining CARD_dict into combi_dict (AMRF-data) in a non-redundant way
for key, values in CARD_dict.items():
    if key not in combi_dict:
        combi_dict[key] = values # Add both key and values if key is not present yet
    else:
        for value in values:
            if value not in combi_dict[key]:
                combi_dict[key].append(value) # If key is present, add only values that are not present in that key

nb_genes = (len(combi_dict))


# STEP 2 : Create a new Excel-file and save the combi_dict
##################################################################################################################################
wb_outfile = xlsxwriter.Workbook(outfile)
ws_outfile = wb_outfile.add_worksheet("AMRF_CARD_combi_reflist")

# Write the header line
header = ["Gene", "Antibiotic"]
ws_outfile.write_row(0, 0, header)

# Write the data to the Excel file
row = 2
for gene, AB_list in combi_dict.items():
    ws_outfile.write(row, 0, gene)
    ws_outfile.write(row, 1, ", ".join(AB_list))
    row += 1

# Close the Excel-file
wb_outfile.close()

# Print a message when the Excel-file has been succesfully created
print(f"Excel-file containing a combined reference list of AMRF & CARD data with {nb_genes} genes has been created at '{outfile}'.")


############ AMRF + CARD + RESF + BN ############ 
# STEP 3 : Create a reference list containing genes of all 4 bioinformatics tools & save in a dictionary
##########################################################################################################################################

# Path to the input file
RESF_BN_list = "/home/guest/BIT11_Traineeship/Ecoli_AMR/Reference_lists/combi_reflist_RESF_BN.xlsx"
# Loading the input file (Excel)
RESF_BN_wb = load_workbook(RESF_BN_list)
# Intializing a dictionary to store genes as keys and antibiotics as values
RESF_BN_dict = {}
# Load the RESF/BN combination reference list into a dictionary
create_dict(RESF_BN_wb, RESF_BN_dict)

# Combining RESF_BN_dict into combi_dict (AMRF-CARD data) in a non-redundant way
for key, values in RESF_BN_dict.items():
    if key not in combi_dict:
        combi_dict[key] = values # Add both key and values if key is not present yet
    else:
        for value in values:
            if value not in combi_dict[key]:
                combi_dict[key].append(value) # If key is present, add only values that are not present in that key

nb_genes_4 = (len(combi_dict))

# STEP 4 : Save the dictionary in an Excel file
#########################################################################################################################
outfile4 = "/home/guest/BIT11_Traineeship/Ecoli_AMR/Reference_lists/combi_all_reflists.xlsx" 
wb_out4 = xlsxwriter.Workbook(outfile4)
ws_out4= wb_out4.add_worksheet("AMRF_CARD_BN_RESF_combi_reflist")

# Write the header line
header = ["Gene", "Antibiotic"]
ws_out4.write_row(0, 0, header)

# Write the data to the Excel file
row = 2
for gene, AB_list in combi_dict.items():
    ws_out4.write(row, 0, gene)
    ws_out4.write(row, 1, ", ".join(AB_list))
    row += 1

# Close the Excel-file
wb_out4.close()

# Print a message when the Excel-file has been succesfully created
print(f"Excel-file containing a combined reference list of BN, RESF, AMRF and CARD data with {nb_genes_4} genes has been created at '{outfile4}'.")                