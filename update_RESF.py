#!/usr/bin/env python
# Script that updates the Excel file with summary of all the bioinformatics tools with the ResFinder results from a summary Excel file.

# Importing the required libraries
import openpyxl

# Read the input & output file
input_file = "/home/guest/BIT11_Traineeship/Ecoli_AMR/ResFinder_tool/ResFinder_summary.xlsx"
output_file = "/home/guest/BIT11_Traineeship/Ecoli_AMR/INFO_MTT_STRAINS.xlsx"

# Load the input and output workbooks
input_wb = openpyxl.load_workbook(input_file)
output_wb = openpyxl.load_workbook(output_file)

# Select the active sheets (assuming there's only one sheet in each file)
input_ws = input_wb.active
output_ws = output_wb.active

# Start row at 3
row = 3

# Loop through each row in the input file containing the ResFinder results for each strain
while row <= input_ws.max_row:
    # Read values from specific cells
    value_2 = input_ws.cell(row=row, column=2).value
    value_3 = input_ws.cell(row=row, column=3).value
    value_4 = input_ws.cell(row=row, column=4).value
    value_5 = input_ws.cell(row=row, column=5).value
    value_6 = input_ws.cell(row=row, column=6).value
    value_7 = input_ws.cell(row=row, column=7).value
    value_8 = input_ws.cell(row=row, column=8).value
    value_9 = input_ws.cell(row=row, column=9).value
    value_10 = input_ws.cell(row=row, column=10).value
    # Row 11 is skipped as it contains results for piperacillin, but this AB is not present on it's own in the output file
    value_12 = input_ws.cell(row=row, column=12).value
    value_13 = input_ws.cell(row=row, column=13).value
    value_14 = input_ws.cell(row=row, column=14).value
    value_15 = input_ws.cell(row=row, column=15).value
    value_16 = input_ws.cell(row=row, column=16).value

    # Write these values to specific cells in the output file
    output_ws.cell(row=row, column=4).value = value_2
    output_ws.cell(row=row, column=11).value = value_3
    output_ws.cell(row=row, column=18).value = value_4
    output_ws.cell(row=row, column=25).value = value_5
    output_ws.cell(row=row, column=32).value = value_6
    output_ws.cell(row=row, column=39).value = value_7
    output_ws.cell(row=row, column=46).value = value_8
    output_ws.cell(row=row, column=53).value = value_9
    output_ws.cell(row=row, column=60).value = value_10
    output_ws.cell(row=row, column=67).value = value_12
    output_ws.cell(row=row, column=74).value = value_13
    output_ws.cell(row=row, column=81).value = value_14
    # Only if row 15 (trimethoprim) & 16 (sulfamethoxazole) are both "R", then write "R" to the output file
    if value_15 == "R" and value_16 == "R":
        output_ws.cell(row=row, column=88).value = "R"
    else:
        output_ws.cell(row=row, column=88).value = "S"
    # Increase the row to move to the next row
    row += 1

# Save the changes to the output Excel file
output_wb.save(output_file)

# Close the workbooks
input_wb.close()
output_wb.close()

print(f"Data has been transferred successfully.")