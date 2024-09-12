#!/usr/bin/env python
# Script to compare genesets found in the data of 284 E. coli WGS analysed with the 4 bioinformatics tools (RESF, BN, RGI-CARD, AMRF+)

# Importing the required libraries
import os
import openpyxl
# Initialize lists to store gene sets per tool
RESF_genes = []
BN_genes = []
RGI_genes = []
AMRF_genes = []

# Specify the input file containing all reference lists & load the workbook
input_file = "/home/guest/BIT11_Traineeship/Ecoli_AMR/Final_documents/reflists.xlsx"
wb = openpyxl.load_workbook(input_file)

# STEP 1: LOAD THE GENES FROM THE REFERENCE LISTS
####################################################################################################################################
# Load the RESF_reflist worksheet
ws_RESF = wb["RESF_reflist"]
for row in ws_RESF.iter_rows(min_row=3, max_row=ws_RESF.max_row, min_col=1, max_col=1):
    for cell in row:
        RESF_genes.append(cell.value.lower())

# Load the BN_reflist worksheet
ws_BN = wb["BN_reflist"]
for row in ws_BN.iter_rows(min_row=3, max_row=ws_BN.max_row, min_col=1, max_col=1):
    for cell in row:
        BN_genes.append(cell.value.lower())

# Load the RGI_reflist worksheet
ws_RGI = wb["CARD_reflist"]
for row in ws_RGI.iter_rows(min_row=3, max_row=ws_RGI.max_row, min_col=1, max_col=1):
    for cell in row:
        RGI_genes.append(cell.value.lower())

# Load the AMRF_reflist worksheet
ws_AMRF = wb["AMRF_reflist"]
for row in ws_AMRF.iter_rows(min_row=3, max_row=ws_AMRF.max_row, min_col=1, max_col=1):
    for cell in row:
        AMRF_genes.append(cell.value.lower())

# Convert the lists to sests
RESF = set(RESF_genes)
BN = set(BN_genes)
RGI = set(RGI_genes)
AMRF = set(AMRF_genes)

# STEP 2 : Calculate the gene counts and shared genes
####################################################################################################################################
# Genes unique to each set
RESF_unique = RESF - (BN | RGI | AMRF)
BN_unique = BN - (RESF | RGI | AMRF)
RGI_unique = RGI - (RESF | BN | AMRF)
AMRF_unique = AMRF - (RESF | BN | RGI)

# Genes shared between two tools
RESF_BN = RESF & BN
RESF_RGI = RESF & RGI
RESF_AMRF = RESF & AMRF
BN_RGI = BN & RGI
BN_AMRF = BN & AMRF
RGI_AMRF = RGI & AMRF

# Genes shared between exactly three tools
RESF_BN_RGI = RESF & BN & RGI
RESF_BN_AMRF = RESF & BN & AMRF
RESF_RGI_AMRF= RESF & RGI & AMRF
BN_RGI_AMRF = BN & RGI & AMRF

# Genes shared across all four tools
intersection = RESF & BN & RGI & AMRF

print(f"Intersection: {len(intersection)} genes occur in all 4 tools.")

# Union = non-redundant/unique genes from all 4 tools combined
all_genes = RESF | BN | RGI | AMRF
"""
RESF_BN_union = RESF | BN
RESF_RGI_union = RESF | RGI
RESF_AMRF_union = RESF | AMRF
BN_RGI_union = BN | RGI
BN_AMRF_union = BN | AMRF
RGI_AMRF_union = RGI | AMRF
"""
print(f"Union: {len(all_genes)} unique genes from all 4 sets combined. \n They are: {all_genes}")

# % genes used by all 4 tools
percentage_shared = len(intersection) / len(all_genes) * 100
#print(f"Percentage of genes used by all 4 tools: {percentage_shared:.2f}%")


# STEP 3 : Write the results to an output file
####################################################################################################################################
output_file = "/home/guest/BIT11_Traineeship/Ecoli_AMR/Final_documents/compare_genesets_3.txt"
with open(output_file, 'w') as f:
    f.write("Amount of genes in each tool:\n")
    f.write("=====================================\n")
    f.write(f"RESF: {len(RESF)} genes\n")
    f.write(f"BN: {len(BN)} genes\n")
    f.write(f"RGI: {len(RGI)} genes\n")
    f.write(f"AMRF: {len(AMRF)} genes\n\n")

    f.write("Unique gene counts per tool:\n")
    f.write("==============================================\n")
    f.write(f"Unique to RESF: {len(RESF_unique)} genes\n")
    f.write(f"Unique to BN: {len(BN_unique)} genes\n")
    f.write(f"Unique to RGI: {len(RGI_unique)} genes\n")
    f.write(f"Unique to AMRF: {len(AMRF_unique)} genes\n\n")

    f.write("Shared gene counts between two tools:\n")
    f.write("=================================\n")
    f.write(f"RESF & BN only: {len(RESF_BN)} genes\n")
    f.write(f"RESF & RGI only: {len(RESF_RGI)} genes\n")
    f.write(f"RESF & AMRF only: {len(RESF_AMRF)} genes\n")
    f.write(f"BN & RGI only: {len(BN_RGI)} genes\n")
    f.write(f"BN & AMRF only: {len(BN_AMRF)} genes\n")
    f.write(f"RGI & AMRF only: {len(RGI_AMRF)} genes\n\n")
    f.write(f"=> {len(intersection)} genes occur in all 4 sets. \n\n")

    f.write("Shared between three tools:\n")
    f.write("=================================\n")
    f.write(f"RESF, BN & RGI only: {len(RESF_BN_RGI)} genes\n")
    f.write(f"RESF, BN & AMRF only: {len(RESF_BN_AMRF)} genes\n")
    f.write(f"RESF, RGI & AMRF only: {len(RESF_RGI_AMRF)} genes\n")
    f.write(f"BN, RGI & AMRF only: {len(BN_RGI_AMRF)} genes\n\n")

    f.write("Shared genes between 4 tools:\n")
    f.write("=================================\n")
    f.write(f"{len(all_genes)} unique genes from all 4 tools combined. \n\n")

    f.write(f"Percentage of genes used by all 4 tools \n")
    f.write(f"{percentage_shared:.2f}%\n")


print(f"Results written to {output_file}")