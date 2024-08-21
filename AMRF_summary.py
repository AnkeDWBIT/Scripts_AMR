#!/usr/bin/env python
# Script that makes an Excel file to summarize the raw data results from AMRFinderPlus on the E. coli WGS

# Importing the required libraries
import os
from os import listdir
from os.path import isfile, join
import re
import csv
import openpyxl
import xlsxwriter

# Folder with AMRF+ results & file with gene-AB reference list 
input_dir = "/home/guest/BIT11_Traineeship/Ecoli_AMR/AMR_output/"
reflist = "/home/guest/BIT11_Traineeship/Ecoli_AMR/combi_reflist_RESF_BN.xlsx"

# List of antibiotics to look for in the ResFinder results
AB_list = ["amikacin", "amoxicillin", "amoxicillin+clavulanic acid", "aztreonam",
            "cefepime", "ceftazidime", "ciprofloxacin", "colistin", "meropenem", 
            "piperacillin", "piperacillin+tazobactam", "tigecycline", "tobramycin",
            "trimethoprim", "sulfamethoxazole"
            ]

# STEP 1 : Make an Excel file to store the summary data further in the script
####################################################################################################################################
output_file = "AMRFPlus_summary.xlsx"
#output_file = "AMRFPlus_summary_blaTEM-1.xlsx"
output_dir = "/home/guest/BIT11_Traineeship/Ecoli_AMR/"
wb = xlsxwriter.Workbook(os.path.join(output_dir, output_file))
ws = wb.add_worksheet("AMRFPlus_summary")

# Write the header line
header = ["File_ID"] + AB_list
ws.write_row(0, 0, header) 

# Initialize the Excel line counter
excel_line = 2

# STEP 2 : Retrieve relevant genes from AMRF+ data by comparing to the reference list and save the associated antibiotic resistances
####################################################################################################################################
# Look in the input directory for the .tsv files
files = [f for f in listdir(input_dir) if isfile(join(input_dir, f))]

# Loop through each file
for file in files:

    # Initialize lists to store genes per sample & antibiotics linked to genes in the sample
    genes = []
    ABs = []

    # Extract the sample name via regular expression (match the first 6 characters = MTT...)
    match = re.match(r'^.{6}', file)
    if match:
        sample = match.group()

    # Open .tsv-files with raw data
    file_path = os.path.join(input_dir, file)
    with open(file_path) as file_to_read:
        next(file_to_read) # Skip the header line when reading the file
        tsv_file = csv.reader(file_to_read, delimiter="\t")
        # Go through each line of data
        for line in tsv_file:
            # Split the gene name on delimiter "_" and take the first part
            gene_name = line[5].split("_")[0]
            #print(gene_name)
            # Save the genename to the list of genes
            genes.append(gene_name)
    
    # Load the reference list workbook & select the combi list (both BN & RESF results)
    ref_wb = openpyxl.load_workbook(reflist)
    ref_ws = ref_wb["combi_reflist"]

    # Loop through all genes from the sample and look for a match in the reference list
    for gene in genes:
        for row in ref_ws.iter_rows(min_row=3, max_col=2, max_row=ref_ws.max_row):
            gene_reflist = row[0].value
            if gene == gene_reflist:
                # Save the antibiotics from the reference list to the list of ABs
                AB_reflist = row[1].value
                antibiotics = [antibiotic.strip() for antibiotic in AB_reflist.split(',')]  # Split the string into individual antibiotics and strip whitespace
                for antibiotic in antibiotics:
                    if antibiotic not in ABs:
                        ABs.append(antibiotic)
                        """
            if gene == "blaTEM-1":
                if "amoxicillin" not in ABs:
                    ABs.append("amoxicillin")
                if "piperacillin" not in ABs:
                    ABs.append("piperacillin")
                    """


    # STEP 3 : Fill in the Excel file with the phenotypic data for each AB in the study for each sample/strain
    #####################################################################################################

    # Make a list to store the phenotypic data and sample name
    phenotypic_data = [sample]
    # Go through the ABs from the study and check if the AMR genes are linked to them
    for AB in AB_list:
        if AB in ABs:
            phenotypic_data.append("R")
        else:
            phenotypic_data.append("S")

    # Write the phenotypic data from the strain to the Excel file
    ws.write_row(excel_line, 0, phenotypic_data)
    # Add to the Excel line counter
    excel_line += 1

# Close the Excel file
wb.close()

# Print a message to indicate the script has finished
print(f"Summary Excel file {output_file} has been created successfully at location {output_dir}!")