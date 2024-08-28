#!/usr/bin/env python
# Script to test the performance of the 4 bioinformatics tools (RESF, BN, CARD, AMRF+) to correctly determine susceptibility compared to MIC (reference)
    # Agreement = correct determination of susceptibility
    # Major errors = tool determines S when MIC is R
    # Minor errors = tool determines R when MIC is S

# Importing the required libraries
import sys
import os
import datetime
import xlsxwriter
import openpyxl

# STEP 1 : HANDLE COMMAND-LINE ARGUMENTS
################################################################################################################################################

# Check if any command-line arguments have been provided
if len(sys.argv) < 3:
    print("Error: Need to provide a command-line argument.")
    print("Usage: python scriptname.py [1] [2][3] [4]")
    print("\t[1] = Full path to input Excel-file with AMR tool comparison results")
    print("\t[2] = Name of the worksheet in the input Excel-file")
    print("\t[3] = Full path where to create the output Excel-file for tool comparison")
    print("\t[4] = Choose a fitting worksheet name for the output Excel-file")

    sys.exit(1)

# Store the command-line argument(s) in an object
input_file = sys.argv[1]
ws_input_name = sys.argv[2]
output_file = sys.argv[3]
ws_name = sys.argv[4]

# STEP 2 : DEFINE FUNCTIONS
################################################################################################################################################
# Functions to convert error from list (found in input file; MIC different from tool) to desired format for the output file
def format_errors(error_list):
    if not error_list:
        return "/"  # Return / if there are no errors
    return f"{len(error_list)}"  # Return the number of errors

# STEP 3 : READ INPUT FILE, CREATE OUTPUT FILE, DEFINE LAYOUT FORMATS & WRITE HEADERS TO THE OUTPUT FILE
################################################################################################################################################
# Load the Excel worksheet
#input_file = "/home/guest/BIT11_Traineeship/Ecoli_AMR/INFO_MTT_STRAINS_updated_RESF_CARD_AMRF_corrected_2.xlsx"
wb = openpyxl.load_workbook(input_file)
ws = wb[ws_input_name]

# Create an Excel output file if it does not exist yet, otherwise just add a new worksheet
#output_file = "/home/guest/BIT11_Traineeship/Ecoli_AMR/tool_performance.xlsx"
wb_output = xlsxwriter.Workbook(output_file)
ws_output = wb_output.add_worksheet(ws_name)

# Define layout formats for the output file
bold_format = wb_output.add_format({'bold': True})
detail_errors = wb_output.add_format({'font_color': '#404040'})
# Apply detail_errors format to the detailed error columns (18-25)
for col in range(18, 26):
    ws_output.set_column(col, col, None, detail_errors)

# Write headers to the output worksheet
ws_output.write(1, 0, "Antibiotic", bold_format)
ws_output.write(0, 1, "RESF", bold_format)
ws_output.write(0, 4, "BN", bold_format)
ws_output.write(0, 7, "CARD", bold_format)
ws_output.write(0, 10, "AMRF+", bold_format)
ws_output.write(1, 1, "% Agreement", bold_format)
ws_output.write(1, 4, "% Agreement", bold_format)
ws_output.write(1, 7, "% Agreement", bold_format)
ws_output.write(1, 10, "% Agreement", bold_format)
ws_output.write(1, 2, "Major errors", bold_format)
ws_output.write(1, 5, "Major errors", bold_format)
ws_output.write(1, 8, "Major errors", bold_format)
ws_output.write(1, 11, "Minor errors", bold_format)
ws_output.write(1, 3, "Minor errors", bold_format)
ws_output.write(1, 6, "Minor errors", bold_format)
ws_output.write(1, 9, "Minor errors", bold_format)
ws_output.write(1, 12, "Minor errors", bold_format)
ws_output.write(16, 0, "Mean agreement per tool", bold_format)
ws_output.write(1, 14, "Mean agreement per AB", bold_format)

# Write headers for detailed description of the errors
ws_output.write(0,18, "RESF")
ws_output.write(1,18, "Major errors")
ws_output.write(1,19, "Minor errors")
ws_output.write(0,20, "BN")
ws_output.write(1,20, "Major errors")
ws_output.write(1,21, "Minor errors")
ws_output.write(0,22, "CARD")
ws_output.write(1,22, "Major errors")
ws_output.write(1,23, "Minor errors")
ws_output.write(0,24, "AMRF+")
ws_output.write(1,24, "Major errors")
ws_output.write(1,25, "Minor errors")

# Add input file, it's used worksheet name & creation date of the output file at the bottom
ws_output.write(20, 18, "Input file:")
ws_output.write(20, 19, input_file)
ws_output.write(21, 18, "Worksheet:")
ws_output.write(21, 19, ws.title)
ws_output.write(22, 18, "Date:")
now = datetime.datetime.now()
current_time_str = now.strftime("%Y-%m-%d %H:%M")
ws_output.write(22, 19, current_time_str)

# STEP 4 : PROCESS INPUT & CALCULATE AGREEMENT PER TOOL PER ANTIBIOTIC
################################################################################################################################################
# Initialize lists to store the agreement percentages per tool
all_RESF_agreement_values = []
all_BN_agreement_values = []
all_CARD_agreement_values = []
all_AMRF_agreement_values = []

# Initialize the row number to start processing data
row_start = 3

# Loop through each antibiotic
for antibiotic_index in range(13):
    # Calculate the starting column indices for MIC, RESF, BN, CARD, AMRF
    base_column = 2 + antibiotic_index * 7
    AB_column = base_column + 0
    MIC_column = base_column + 1
    RESF_column = base_column + 2
    BN_column = base_column + 3
    CARD_column = base_column + 4
    AMRF_column = base_column + 5

    # Initialize counters & dictionaries to store the results
    RESF_agreement = 0
    RESF_major_errors = []
    RESF_minor_errors = []

    BN_agreement = 0
    BN_major_errors = []
    BN_minor_errors = []

    CARD_agreement = 0
    CARD_major_errors = []
    CARD_minor_errors = []

    AMRF_agreement = 0
    AMRF_major_errors = []
    AMRF_minor_errors = []

    row = row_start  # Reset row for each antibiotic

    # Go through each row in the Excel file
    while row <= ws.max_row:
        # Read values from specific cells
        strain = ws.cell(row=row, column=1).value
        AB = ws.cell(row=1, column=AB_column).value  # This assumes the antibiotic name is in the first row
        MIC = ws.cell(row=row, column=MIC_column).value
        RESF = ws.cell(row=row, column=RESF_column).value
        BN = ws.cell(row=row, column=BN_column).value
        CARD = ws.cell(row=row, column=CARD_column).value
        AMRF = ws.cell(row=row, column=AMRF_column).value
        #print(strain, AB, MIC, RESF, BN, CARD, AMRF)

        # Calculate the agreement & errors per tool
        if RESF == MIC:
            RESF_agreement += 1
        else:
            if MIC == "R" and RESF == "S":
                RESF_major_errors.append(strain)
            if MIC == "S" and RESF == "R":
                RESF_minor_errors.append(strain)

        if BN == MIC:
            BN_agreement += 1
        else:
            if MIC == "R" and BN == "S":
                BN_major_errors.append(strain)
            if MIC == "S" and BN == "R":
                BN_minor_errors.append(strain)
        
        if CARD == MIC:
            CARD_agreement += 1
        else:
            if MIC == "R" and CARD == "S":
                CARD_major_errors.append(strain)
            if MIC == "S" and CARD == "R":
                CARD_minor_errors.append(strain)

        if AMRF == MIC:
            AMRF_agreement += 1
        else:
            if MIC == "R" and AMRF == "S":
                AMRF_major_errors.append(strain)
            if MIC == "S" and AMRF == "R":
                AMRF_minor_errors.append(strain)

        # Increase the row to move to the next row
        row += 1
   
    # Calculate the agreement percentage & show the errors per tool 
    total_strains = ws.max_row - row_start + 1
    RESF_agreement_percentage = RESF_agreement / total_strains * 100
    BN_agreement_percentage = BN_agreement / total_strains * 100
    CARD_agreement_percentage = CARD_agreement / total_strains * 100
    AMRF_agreement_percentage = AMRF_agreement / total_strains * 100


    # Calculate the mean agreement per antibiotic
    mean_agreement_per_AB = (RESF_agreement_percentage + BN_agreement_percentage + CARD_agreement_percentage + AMRF_agreement_percentage) / 4

    # Add agreement percentage per tool to a list, so that after looping through all antibiotics, the mean agreement per tool can be calculated
    all_RESF_agreement_values.append(RESF_agreement_percentage)
    all_BN_agreement_values.append(BN_agreement_percentage)
    all_CARD_agreement_values.append(CARD_agreement_percentage)
    all_AMRF_agreement_values.append(AMRF_agreement_percentage)

# STEP 5 : WRITE RESULTS TO THE OUTPUT FILE
################################################################################################################################################
    # Add the results to the output file
    ws_output.write(2 + antibiotic_index, 0, AB, bold_format)
    ws_output.write(2 + antibiotic_index, 1, round(RESF_agreement_percentage, 2))
    ws_output.write(2 + antibiotic_index, 2, format_errors(RESF_major_errors))
    ws_output.write(2 + antibiotic_index, 3, format_errors(RESF_minor_errors))
    ws_output.write(2 + antibiotic_index, 4, round(BN_agreement_percentage, 2))
    ws_output.write(2 + antibiotic_index, 5, format_errors(BN_major_errors))
    ws_output.write(2 + antibiotic_index, 6, format_errors(BN_minor_errors))
    ws_output.write(2 + antibiotic_index, 7, round(CARD_agreement_percentage, 2))
    ws_output.write(2 + antibiotic_index, 8, format_errors(CARD_major_errors))
    ws_output.write(2 + antibiotic_index, 9, format_errors(CARD_minor_errors))
    ws_output.write(2 + antibiotic_index, 10, round(AMRF_agreement_percentage, 2))
    ws_output.write(2 + antibiotic_index, 11, format_errors(AMRF_major_errors))
    ws_output.write(2 + antibiotic_index, 12, format_errors(AMRF_minor_errors))
    ws_output.write(2 + antibiotic_index, 14, round(mean_agreement_per_AB, 2))

    # Add the detailed errors to the output file, first convert the lists to strings
    RESF_major_errors = ", ".join(RESF_major_errors)
    RESF_minor_errors = ", ".join(RESF_minor_errors)
    BN_major_errors = ", ".join(BN_major_errors)
    BN_minor_errors = ", ".join(BN_minor_errors)
    CARD_major_errors = ", ".join(CARD_major_errors)
    CARD_minor_errors = ", ".join(CARD_minor_errors)
    AMRF_major_errors = ", ".join(AMRF_major_errors)
    AMRF_minor_errors = ", ".join(AMRF_minor_errors)

    ws_output.write(2 + antibiotic_index, 18, RESF_major_errors)
    ws_output.write(2 + antibiotic_index, 19, RESF_minor_errors)
    ws_output.write(2 + antibiotic_index, 20, BN_major_errors)
    ws_output.write(2 + antibiotic_index, 21, BN_minor_errors)
    ws_output.write(2 + antibiotic_index, 22, CARD_major_errors)
    ws_output.write(2 + antibiotic_index, 23, CARD_minor_errors)
    ws_output.write(2 + antibiotic_index, 24, AMRF_major_errors)
    ws_output.write(2 + antibiotic_index, 25, AMRF_minor_errors)

# Calculate mean % agreement of each tool (can only be done after looping through all antibiotics)
mean_RESF_agreement = sum(all_RESF_agreement_values) / 13
mean_BN_agreement = sum(all_BN_agreement_values) / 13
mean_CARD_agreement = sum(all_CARD_agreement_values) / 13
mean_AMRF_agreement = sum(all_AMRF_agreement_values) / 13

# Add the mean agreement per tool to the output file
ws_output.write(16, 1, round(mean_RESF_agreement, 2))
ws_output.write(16, 4, round(mean_BN_agreement, 2))
ws_output.write(16, 7, round(mean_CARD_agreement, 2))
ws_output.write(16, 10, round(mean_AMRF_agreement, 2))

# Save the changes to the output Excel file
wb_output.close()

# Print a message when the script has finished
print(f"Tool performance has been calculated and saved to {output_file}.")