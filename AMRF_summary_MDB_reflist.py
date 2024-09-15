#!/usr/bin/env python
# Script that makes an Excel file to summarize the raw data results from AMRFinderPlus on the E. coli WGS

# Importing the required libraries
import os
from os import listdir
from os.path import isfile, join
import re
import csv
import openpyxl
import xlsxwriter

# Folder with AMRF+ results & file with gene-AB reference list 
input_dir = "/home/guest/BIT11_Traineeship/Ecoli_AMR/Ruwe_data/AMR_output/"
reflist = "/home/guest/BIT11_Traineeship/Ecoli_AMR/Final_documents/reflists.xlsx"

# List of antibiotics to look for in the ResFinder results
AB_list = ["amikacin", "amoxicillin", "amoxicillin+clavulanic acid", "aztreonam",
            "cefepime", "ceftazidime", "ciprofloxacin", "colistin", "meropenem", 
            "piperacillin", "piperacillin+tazobactam", "tigecycline", "tobramycin",
            "trimethoprim", "sulfamethoxazole"
            ]

# Open the reference list with vocabulary adjustments
adjust_vocab = openpyxl.load_workbook("/home/guest/BIT11_Traineeship/Ecoli_AMR/Final_documents/reflists_MDB.xlsx")
adjust_ws = adjust_vocab["AMRF_CARD_BN_RESF_combi_reflist"]

# STEP 1 : Make an Excel file to store the summary data further in the script
####################################################################################################################################
#output_file = "AMRFPlus_summary.xlsx"
output_file = "AMRF_summary_reflist_MDB_test.xlsx"
output_dir = "/home/guest/BIT11_Traineeship/Ecoli_AMR/Final_documents/Summary_Excel/"
wb = xlsxwriter.Workbook(os.path.join(output_dir, output_file))
ws = wb.add_worksheet("AMRFPlus_summary")

# Write the header line
header = ["File_ID"] + AB_list
ws.write_row(0, 0, header) 

# Initialize the Excel line counter
excel_line = 2

# STEP 2 : Retrieve relevant genes from AMRF+ data by comparing to the reference list and save the associated antibiotic resistances
####################################################################################################################################
# Look in the input directory for the .tsv files
files = [f for f in listdir(input_dir) if isfile(join(input_dir, f))]

# Loop through each file
for file in files:

    # Initialize lists to store genes per sample & antibiotics linked to genes in the sample
    genes = []
    ABs = []

    # Extract the sample name via regular expression (match the first 6 characters = MTT...)
    match = re.match(r'^.{6}', file)
    if match:
        sample = match.group()

    # Open .tsv-files with raw data
    file_path = os.path.join(input_dir, file)
    with open(file_path) as file_to_read:
        next(file_to_read) # Skip the header line when reading the file
        tsv_file = csv.reader(file_to_read, delimiter="\t")
        # Go through each line of data
        for line in tsv_file:
            # Split the gene name on delimiter "_" and take the first part
            gene_name = line[5].split("_")[0]
            #print(gene_name)
            # Save the genename to the list of genes
            genes.append(gene_name)
    
    # Load the reference list workbook & select the combi list (both BN & RESF results)
    ref_wb = openpyxl.load_workbook(reflist)
    ref_ws = ref_wb["RESF_BN_combi_reflist"]

    # Loop through all genes from the sample and look for a match in the reference list
    for gene in genes:
        if gene =="AAC(3)-IVa":
            gene = "aac(3)-IV"
        if gene =="AAC(3)-IVa":
            gene = "aac(3)-IV"
        if gene == "AAC(3)-VIa":
            gene = "aac(3)-VIa"
        if gene == "aadA":
            gene = "aadA1"
        if gene == "ant(3'')-Ia":
            gene = "aadA1"
        if gene == "blaACT":
            gene = "blaACT-1"
        if gene == "blaCTX-M":
            gene = "blaCTX-M-1"
        if gene == "blaTEM":
            gene = "blaTEM-1A"
        if gene == "blaTEMp":
            gene = "blaTEM-1A"
        if gene == "catI":
            gene = "catA1"
        if gene == "CTX-M-1":
            gene = "blaCTX-M-1"
        if gene == "CTX-M-15":
            gene = "blaCTX-M-15"
        if gene == "DfrA36":
            gene = "dfrA36"
        if gene == "Enterobacter cloacae acrA":
            gene = "acrA"
        if gene == "ErmB":
            gene = "erm(B)"
        if gene == "Escherichia coli acrA":
            gene = "acrA"
        if gene == "Escherichia coli ampC beta-lactamase":
            gene = "blaACT-1"
        if gene == "Escherichia coli cyaA with mutation conferring resistance to fosfomycin":
            gene = "cyaA"
        if gene == "Escherichia coli GlpT with mutation conferring resistance to fosfomycin":
            gene = "glpT"
        if gene == "Escherichia coli gyrA conferring resistance to fluoroquinolones":
            gene = "gyrA"
        if gene == "Escherichia coli gyrA with mutation conferring resistance to triclosan":
            gene = "gyrA"
        if gene == "Escherichia coli mdfA":
            gene = "mdf(A)"
        if gene == "Escherichia coli nfsA mutations conferring resistance to nitrofurantoin":
            gene = "nfsA"
        if gene == "Escherichia coli parC conferring resistance to fluoroquinolones":
            gene = "parC"
        if gene == "Escherichia coli PtsI with mutation conferring resistance to fosfomycin":
            gene = "ptsI"
        if gene == "Escherichia coli soxR with mutation conferring antibiotic resistance":
            gene = "soxR"
        if gene == "Escherichia coli UhpT with mutation conferring resistance to fosfomycin":
            gene = "uhpT"
        if gene == "mdf(A)":
            gene = "mdf(A)"
        if gene == "mphA":
            gene = "mph(A)"
        if gene == "mphB":
            gene = "mph(B)"
        if gene == "OXA-1":
            gene = "blaOXA-1"
        if gene == "OXA-2":
            gene = "blaOXA-2"
        if gene == "OXA-9":
            gene = "blaOXA-9"
        if gene == "qnrE":
            gene = "qnrE1"
        if gene == "qnrS":
            gene = "qnrS1"
        if gene == "QnrS1":
            gene = "qnrS1"
        if gene == "Salmonella enterica gyrA with mutation conferring resistance to triclosan":
            gene = "gyrA"
        if gene == "SAT-2":
            gene = "sat2"
        if gene == "SHV-1":
            gene = "blaSHV-1"
        if gene == "SHV-4":
            gene = "blaSHV-4"
        if gene == "TEM-1":
            gene = "blaTEM-1"
        if gene == "TEM-214":
            gene = "blaTEM-214"
        if gene == "TEM-24":
            gene = "blaTEM-24"
        if gene == "TEM-30":
            gene = "blaTEM-30"
        if gene == "TEM-32":
            gene = "blaTEM-32"
        if gene == "TEM-33":
            gene = "blaTEM-33"
        if gene == "TEM-34":
            gene = "blaTEM-34"
        if gene == "TEM-40":
            gene = "blaTEM-40"
        if gene == "TEM-52":
            gene = "blaTEM-52"
        if gene == "TEM-54":
            gene = "blaTEM-54"
        
        for row in ref_ws.iter_rows(min_row=3, max_col=2, max_row=ref_ws.max_row):
            gene_reflist = row[0].value
            if gene == gene_reflist:
                #print(f"Final gene used for AB lookup: {gene}")  # Debug to check if adjusted name is used
                # Save the antibiotics from the reference list to the list of ABs
                AB_reflist = row[1].value
                antibiotics = [antibiotic.strip() for antibiotic in AB_reflist.split(',')]  # Split the string into individual antibiotics and strip whitespace
                for antibiotic in antibiotics:
                    if antibiotic not in ABs:
                        ABs.append(antibiotic) 
        



    # STEP 3 : Fill in the Excel file with the phenotypic data for each AB in the study for each sample/strain
    #####################################################################################################

    # Make a list to store the phenotypic data and sample name
    phenotypic_data = [sample]
    # Go through the ABs from the study and check if the AMR genes are linked to them
    for AB in AB_list:
        if AB in ABs:
            phenotypic_data.append("R")
        else:
            phenotypic_data.append("S")

    # Write the phenotypic data from the strain to the Excel file
    ws.write_row(excel_line, 0, phenotypic_data)
    # Add to the Excel line counter
    excel_line += 1

# Close the Excel file
wb.close()

# Print a message to indicate the script has finished
print(f"Summary Excel file {output_file} has been created successfully at location {output_dir}!")