#!/usr/bin/env python
# Script that updates the Excel file with summary of all the bioinformatics tools with the AMRF+ & CARD results from the summary Excel files.

# Importing the required libraries
import openpyxl

# Read the input & output files
AMRF_input = "/home/guest/BIT11_Traineeship/Ecoli_AMR/AMRFPlus_summary.xlsx"
CARD_input = "/home/guest/BIT11_Traineeship/Ecoli_AMR/CARD_summary.xlsx"
output_file = "/home/guest/BIT11_Traineeship/Ecoli_AMR/INFO_MTT_STRAINS_updated_RESF_CARD_AMRF.xlsx"

# Load the input and output workbooks
AMRF_wb = openpyxl.load_workbook(AMRF_input)
CARD_wb = openpyxl.load_workbook(CARD_input)
output_wb = openpyxl.load_workbook(output_file)

# Select the active sheets (assuming there's only one sheet in each file)
AMRF_ws = AMRF_wb.active
CARD_ws = CARD_wb.active
output_ws = output_wb["Comparative AMR (2)"]

# STEP 1 : Transfer the AMRF+ results to the output file
################################################################################################################################################
# Loop through each row in the input file containing the AMRF+ results for each strain
# Start row at 3
row = 3
while row <= AMRF_ws.max_row:
    # Read values from specific cells
    value_2 = AMRF_ws.cell(row=row, column=2).value
    value_3 = AMRF_ws.cell(row=row, column=3).value
    value_4 = AMRF_ws.cell(row=row, column=4).value
    value_5 = AMRF_ws.cell(row=row, column=5).value
    value_6 = AMRF_ws.cell(row=row, column=6).value
    value_7 = AMRF_ws.cell(row=row, column=7).value
    value_8 = AMRF_ws.cell(row=row, column=8).value
    value_9 = AMRF_ws.cell(row=row, column=9).value
    value_10 = AMRF_ws.cell(row=row, column=10).value
    # Row 11 is skipped as it contains results for piperacillin, but this AB is not present on it's own in the output file
    value_12 = AMRF_ws.cell(row=row, column=12).value
    value_13 = AMRF_ws.cell(row=row, column=13).value
    value_14 = AMRF_ws.cell(row=row, column=14).value
    value_15 = AMRF_ws.cell(row=row, column=15).value
    value_16 = AMRF_ws.cell(row=row, column=16).value

    # Write these values to specific cells in the output file
    output_ws.cell(row=row, column=7).value = value_2
    output_ws.cell(row=row, column=14).value = value_3
    output_ws.cell(row=row, column=21).value = value_4
    output_ws.cell(row=row, column=28).value = value_5
    output_ws.cell(row=row, column=35).value = value_6
    output_ws.cell(row=row, column=42).value = value_7
    output_ws.cell(row=row, column=49).value = value_8
    output_ws.cell(row=row, column=56).value = value_9
    output_ws.cell(row=row, column=63).value = value_10
    output_ws.cell(row=row, column=70).value = value_12
    output_ws.cell(row=row, column=77).value = value_13
    output_ws.cell(row=row, column=84).value = value_14
    # Only if row 15 (trimethoprim) & 16 (sulfamethoxazole) are both "R", then write "R" to the output file
    if value_15 == "R" and value_16 == "R":
        output_ws.cell(row=row, column=91).value = "R"
    else:
        output_ws.cell(row=row, column=91).value = "S"
    # Increase the row to move to the next row
    row += 1

# STEP 2 : Transfer the CARD results to the output file
################################################################################################################################################
# Loop through each row in the input file containing the CARD results for each strain
# Start row at 3
row = 3
while row <= AMRF_ws.max_row:
    # Read values from specific cells
    val_2 = CARD_ws.cell(row=row, column=2).value
    val_3 = CARD_ws.cell(row=row, column=3).value
    val_4 = CARD_ws.cell(row=row, column=4).value
    val_5 = CARD_ws.cell(row=row, column=5).value
    val_6 = CARD_ws.cell(row=row, column=6).value
    val_7 = CARD_ws.cell(row=row, column=7).value
    val_8 = CARD_ws.cell(row=row, column=8).value
    val_9 = CARD_ws.cell(row=row, column=9).value
    val_10 = CARD_ws.cell(row=row, column=10).value
    # Row 11 is skipped as it contains results for piperacillin, but this AB is not present on it's own in the output file
    val_12 = CARD_ws.cell(row=row, column=12).value
    val_13 = CARD_ws.cell(row=row, column=13).value
    val_14 = CARD_ws.cell(row=row, column=14).value
    val_15 = CARD_ws.cell(row=row, column=15).value
    val_16 = CARD_ws.cell(row=row, column=16).value

    # Write these values to specific cells in the output file
    output_ws.cell(row=row, column=6).value = val_2
    output_ws.cell(row=row, column=13).value = val_3
    output_ws.cell(row=row, column=20).value = val_4
    output_ws.cell(row=row, column=27).value = val_5
    output_ws.cell(row=row, column=34).value = val_6
    output_ws.cell(row=row, column=41).value = val_7
    output_ws.cell(row=row, column=48).value = val_8
    output_ws.cell(row=row, column=58).value = val_9
    output_ws.cell(row=row, column=62).value = val_10
    output_ws.cell(row=row, column=69).value = val_12
    output_ws.cell(row=row, column=76).value = val_13
    output_ws.cell(row=row, column=83).value = val_14
    # Only if row 15 (trimethoprim) & 16 (sulfamethoxazole) are both "R", then write "R" to the output file
    if val_15 == "R" and val_16 == "R":
        output_ws.cell(row=row, column=90).value = "R"
    else:
        output_ws.cell(row=row, column=90).value = "S"
    # Increase the row to move to the next row
    row += 1


# Save the changes to the output Excel file
output_wb.save(output_file)

# Close the workbooks
AMRF_wb.close()
CARD_wb.close()
output_wb.close()

# Print a message to the user
print(f"Data has been transferred successfully from {AMRF_input} and {CARD_input} to {output_file}.")
