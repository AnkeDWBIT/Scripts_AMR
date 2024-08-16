#!/usr/bin/env python
# Script to combine the RESF & BN reference lists containing antibiotic resistance linked to genes into one reference list

# Importing the required libraries
from openpyxl import load_workbook
import xlsxwriter

# Saving the paths to the files in an object
BN_list = "/home/guest/BIT11_Traineeship/Ecoli_AMR/BN_reflist.xlsx"
RESF_list = "/home/guest/BIT11_Traineeship/Ecoli_AMR/RESF_reflist.xlsx" 
outfile = "/home/guest/BIT11_Traineeship/Ecoli_AMR/combi_reflist_RESF_BN.xlsx" 

# Loading input-& output-file
BN_wb = load_workbook(BN_list)
RESF_wb = load_workbook(RESF_list)

# STEP 1 : Combining data from BN- & RESF-reference lists in a dictionary
###########################################################################################################################

# Initializing dictionaries to store the data
combi_dict = {}
BN_dict = {}

# Function to create a dictionary from a workbook's first sheet
def create_dict(wb,dict):
    ws = wb.active
    for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        key = row[0]
        value = row[1].lower()
        antibiotics = [antibiotic.strip() for antibiotic in value.split(',')]  # Split the string into individual antibiotics and strip whitespace
        dict[key] = antibiotics

# Storing the data from combi_wb & BN_wb in a dictionary each
create_dict(RESF_wb, combi_dict)
create_dict(BN_wb, BN_dict)
#print(combi_dict)

# Combining BN_dict into combi-dict in a non-redundant way
for key, values in BN_dict.items():
    if key not in combi_dict:
        combi_dict[key] = values # Add both key and values if key is not present yet
    else:
        for value in values:
            if value not in combi_dict[key]:
                combi_dict[key].append(value) # If key is present, add only values that are not present in that key

# Print the combi_dict (should have 120 keys because there are 120 unique genes between the two reference lists)
#print(combi_dict)
nb_genes = (len(combi_dict))


# STEP 2 : Create a new Excel-file and save the combi_dict
##################################################################################################################################
wb_outfile = xlsxwriter.Workbook(outfile)
ws_outfile = wb_outfile.add_worksheet("combi_reflist")

# Write the header line
header = ["Gene", "Antibiotic"]
ws_outfile.write_row(0, 0, header)

# Write the data to the Excel file
row = 2
for gene, AB_list in combi_dict.items():
    ws_outfile.write(row, 0, gene)
    ws_outfile.write(row, 1, ", ".join(AB_list))
    row += 1

# Close the Excel-file
wb_outfile.close()

# Print a message when the Excel-file has been succesfully created
print(f"Excel-file containing a combined reference list of BN & RESF data with {nb_genes} genes has been created at '{outfile}'.")